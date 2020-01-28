

import xlrd
from xlrd import open_workbook
import os
import shutil
import subprocess
import csv
import openpyxl
from openpyxl.reader.excel import load_workbook
import locale
import datetime
import PyPDF2
from PyPDF2 import PdfFileReader, PdfFileWriter
from PyPDF2 import PdfFileMerger, PageRange

global ws

import pandas as pd
import re
import glob
from glob import glob
import PyPDF2
import datetime
import fpdf
from fpdf import FPDF
import io
import pytesseract
# from pytesseract import image_to_string
# from pytesseract import image_to_boxes
import PIL
from PIL import Image as ImagePIL
from io import StringIO
import wand
from wand.image import Image
import cv2
import numpy as np
from PIL import Image as im
import scipy
from scipy.ndimage import interpolation as inter

# Camlot

# import camelot

import PdfminerCheck as pdfminer

import spacy

from spacy.lang.en import STOP_WORDS
# from spacy.pipeline import EntityRuler
# from spacy.lang.en import English
# from spacy.matcher import PhraseMatcher
from spacy.tokens import Span

import json


import TabulaExtract

from custom_modules.IR56B_Form import customDataCleaningForIR56B

cwd = os.getcwd()
workflow_Excel_path = os.path.join(cwd, "Config.xlsx")
wb = load_workbook(workflow_Excel_path)
workbook = open_workbook(workflow_Excel_path)
sheet = workbook.sheet_by_index(0)
ghostscriptpath = sheet.cell_value(6, 1)
imagemagickpath = sheet.cell_value(7, 1)

path_to_gs = ghostscriptpath
os.environ['PATH'] += os.pathsep + path_to_gs
os.environ['PATH'] += os.pathsep + imagemagickpath
# import StanfordNER


# >>>>>>>------   Support Functions ----------------------------

# Current Working Directory
cwd = os.getcwd()
# print(cwd)
# workflow_Excel_path = os.path.join(cwd, "Config.xlsx")

tessdata_dir_config = '--tessdata-dir' + '"' + os.path.join(cwd, "Tesseract-OCR-4.0\\tessdata") + '"';
pytesseract.pytesseract.tesseract_cmd = os.path.join(cwd, "Tesseract-OCR-4.0\\tesseract.exe")



spacylgmodelpath = os.path.join(cwd, "en_core_web_sm-2.1.0-custom")

# This is for custom trained model
spacyCustomModelpath = os.path.join(cwd, "en_core_web_sm-2.1.0-custom")

# print(spacylgmodelpath)

_nlp = spacy.load(spacylgmodelpath)




# --------------- Str cleanup in Panda list -----
def strcleanup(strinp):
    return str(strinp).replace("[", "").replace("]", "").replace("'", "")


def amountFormatter(Amnt):
    locale.setlocale(locale.LC_ALL, 'en-US')
    amountRounded = locale.format("%d", Amnt, grouping=True)
    return amountRounded


# ----  Get Source path details ------->>>

def getSourcePathDetails():
    workbook = open_workbook(workflow_Excel_path)
    sheet = workbook.sheet_by_index(0)
    # if sheet.cell_value(0, 2) == "Applicable":
    return sheet.cell_value(0, 1)
    # elif sheet.cell_value(1, 2) == "Applicable":


#         return sheet.cell_value(1, 1)


def getResultPathDetails():
    workbook = open_workbook(workflow_Excel_path)
    sheet = workbook.sheet_by_index(0)
    # if sheet.cell_value(1, 2) == "Applicable":
    return sheet.cell_value(1, 1)


# -------------------------------------
# ---  Preferences ------------------


def getPrefSplitpageByLogo():
    workbook = open_workbook(workflow_Excel_path)
    sheet = workbook.sheet_by_index(0)
    if sheet.cell_value(9, 2) == "Applicable":
        return sheet.cell_value(9, 2)


def getPagePreferences():
    workbook = open_workbook(workflow_Excel_path)
    sheet = workbook.sheet_by_index(0)
    if sheet.cell_value(8, 2) == "Applicable":
        return sheet.cell_value(8, 1)


def getExtractDataPref():
    workbook = open_workbook(workflow_Excel_path)
    sheet = workbook.sheet_by_index(0)
    return True


def getExtractTablesPref():
    workbook = open_workbook(workflow_Excel_path)
    sheet = workbook.sheet_by_index(0)
    if sheet.cell_value(11, 2) == "Applicable":
        return True


# --- End of Preferences ------------------


def makddirectory(name):
    current_directory = os.getcwd()
    final_directory = os.path.join(current_directory, name)
    if not os.path.exists(final_directory):
        # print("***************final_directory********************")
        # print(final_directory)
        os.makedirs(final_directory)


def makdsubdirectory(dirname, subdirname):
    current_directory = os.getcwd()
    dirname = os.path.join(current_directory, dirname)
    final_directory = os.path.join(dirname, subdirname)
    # if not os.path.exists(final_directory):
    # print("***************final_directory in sub directory ********************")
    # print(final_directory)
    os.makedirs(final_directory)


# To check the image clarity
def variance_of_laplacian(image):
    # compute the Laplacian of the image and then return the focus
    # measure, which is simply the variance of the Laplacian
    return cv2.Laplacian(image, cv2.CV_64F).var()


# BINARY_THREHOLD = 180


def remove_noise_and_smooth(file_name):
    # print("LLLLLLLLLLL-----")
    # print(file_name)
    #    img = cv2.imread(file_name, 0)
    #
    #    img = cv2.resize(img, None, fx=1.5, fy=1.5, interpolation=cv2.INTER_CUBIC)
    #
    #    # Apply dilation and erosion to remove some noise
    #    kernel = np.ones((1, 1), np.uint8)
    #    img = cv2.dilate(img, kernel, iterations=1)
    #    img = cv2.erode(img, kernel, iterations=1)
    #
    #    # Apply blur to smooth out the edges
    #    #img = cv2.GaussianBlur(img, (5, 5), 0)
    #    img3 = cv2.bilateralFilter(img,9,75,75)
    #
    #    # Apply threshold to get image with only b&w (binarization)
    #    ret, filtered = cv2.threshold(img3, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    #
    #    resimg = filtered
    #    return resimg
    #
    imgforlp = cv2.imread(file_name, 0)
    laplacian_score = variance_of_laplacian(imgforlp)
    im3 = ImagePIL.open(file_name)

    # if laplacian_score < 2500:
    # white = im3.filter(ImageFilter.BLUR).filter(ImageFilter.MaxFilter(15))
    # im3.save(file_name, dpi=(400,400))
    # else:
    # im3.save(file_name, dpi=(300,300))

    img = cv2.imread(file_name, 0)

    if laplacian_score > 5500:

        # print("Laplacian value --->")
        # print(laplacian_score)
        # Added by Mani - Get accurate numbers and slashes

        imgpreprocess = cv2.adaptiveThreshold(img.astype(np.uint8), 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                              cv2.THRESH_BINARY, 31, 2)

        # img555 = cv2.threshold(img555.astype(np.uint8),127,255,cv2.THRESH_BINARY)
        height, width = imgpreprocess.shape[:2]

        imgpreprocess = cv2.resize(imgpreprocess, None, fx=1.5, fy=1.5, interpolation=cv2.INTER_CUBIC)

        kernel = np.ones((3, 3), np.uint8)
        imgpreprocess = cv2.dilate(imgpreprocess, kernel, iterations=2)
        imgpreprocess = cv2.erode(imgpreprocess, kernel, iterations=1)

        imgpreprocess = cv2.bilateralFilter(imgpreprocess, 9, 75, 75)

        ##img = cv2.cvtColor(img555, cv2.COLOR_BGR2GRAY)

        # Apply threshold to get image with only b&w (binarization)
        ret, filtered = cv2.threshold(imgpreprocess, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
        resimg = filtered

        resimg = imgpreprocess

    else:

        # print(qualityscore)
        img = cv2.resize(img, None, fx=1.5, fy=1.5, interpolation=cv2.INTER_CUBIC)
        # Apply dilation and erosion to remove some noise
        kernel = np.ones((1, 1), np.uint8)
        img = cv2.dilate(img, kernel, iterations=1)
        img = cv2.erode(img, kernel, iterations=1)

        # Apply blur to smooth out the edges
        # img = cv2.GaussianBlur(img, (5, 5), 0)
        img3 = cv2.bilateralFilter(img, 9, 75, 75)

        # Apply threshold to get image with only b&w (binarization)
        ret, filtered = cv2.threshold(img3, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
        resimg = filtered

    return resimg


# IMAGE_SIZE = 2300

def set_image_dpi(file_path, index, file):
    #    im = ImagePIL.open(file_path)
    #    length_x, width_y = im.size
    #    factor = max(1, int(IMAGE_SIZE / length_x))
    #    size = factor * length_x, factor * width_y
    #    # size = (1800, 1800)
    #    im_resized = im.resize(size, ImagePIL.ANTIALIAS)
    #    #img3 = cv2.imread(file_path)
    #
    #    #im_resized = cv2.resize(img3, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    #
    #    im_resized.save(os.path.join("DataProcess", file)+str(index)+"-temp-first-pagewihtdpi.png", dpi=(600, 600))
    #
    #    savedpath = os.path.join("DataProcess", file)+str(index)+"-temp-first-pagewihtdpi.png"
    #
    savedpath = file_path

    # savedimage = cv2.imread(savedpath)
    # img = cv2.resize(savedimage, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    # Dilate to improve OCR results
    # img = cv2.imread(savedpath)
    # kernel = np.ones((1, 1), np.uint8)
    # im_dilated = cv2.dilate(img, kernel, iterations=2)
    # cv2.imwrite(os.path.join("PDFs", file)+str(index)+"-temp-first-pagewihtdpi.png",im_dilated)

    return savedpath


def skewCorrection(file_path, index, file):
    img = ImagePIL.open(file_path)

    # convert to binary
    wd, ht = img.size
    pix = np.array(img.convert('1').getdata(), np.uint8)
    bin_img = 1 - (pix.reshape((ht, wd)) / 255.0)
    plt.imshow(bin_img, cmap='gray')
    plt.savefig('binary.png')

    def find_score(arr, angle):
        data = inter.rotate(arr, angle, reshape=False, order=0)
        hist = np.sum(data, axis=1)
        score = np.sum((hist[1:] - hist[:-1]) ** 2)
        return hist, score

    delta = 1
    limit = 5
    angles = np.arange(-limit, limit + delta, delta)
    scores = []
    for angle in angles:
        hist, score = find_score(bin_img, angle)
        scores.append(score)

    best_score = max(scores)
    best_angle = angles[scores.index(best_score)]
    # print('Best angle: {}'.format(best_angle))

    # correct skew
    data = inter.rotate(bin_img, best_angle + 0.25, reshape=False, order=0)
    img = im.fromarray((255 * data).astype("uint8")).convert("RGB")
    # img.save('C:\\Users\\manivel.muthusamy\\DataExtract\\Cognitive-RPA-PDF-data-extract-by-ocr\\PDFs\\test\\skew_corrected.png')

    img.save(os.path.join(cwd + "\\DataProcess_56B", file) + str(index) + "-temp-first-pagewithdpi-skewed.png",
             dpi=(600, 600))

    savedpath = os.path.join(cwd + "\\DataProcess_56B", file) + str(index) + "-temp-first-pagewithdpi-skewed.png"
    return savedpath


# >>>>>------   Support Functions ----------------------------


# --------------- Generate new files list to be processed  -----

def processDataExtraction_56B():

    try:
        shutil.rmtree(os.path.join(cwd, "DataProcess_56B"))
    except Exception as e:
        print(e)
        print("Exception during delete dataprocess")
    # Get path from Excel settings (settings tab)
    rootPath = getSourcePathDetails()

    #
    # print("rootPath -- " + rootPath)

    pathforclientdirs = os.path.join(rootPath)

    resultPath = getResultPathDetails()

    if getPrefSplitpageByLogo():
        splitpagesbyAssignees(pathforclientdirs, "PDFs", resultPath)
        shutil.rmtree(os.path.join(cwd, "PDFs"))

    Sheet3 = wb['Results_56B']

    # ----  To remove all existing rows generated previously ------

    # for i in range(2, Sheet3.max_row + 1):
    # if Sheet2.cell(row=i, column=1).value != " " or Sheet2.cell(row=i, column=1).value != None:
    # Sheet3.delete_rows(i,1)

    for rc in range(2, Sheet3.max_row + 1):
        # if Sheet2.cell(row=i, column=1).value != " " or Sheet2.cell(row=i, column=1).value != None:
        # print("--------------  Sheet3.max_row + 1 ==============-----")
        # print(rc)
        Sheet3.delete_rows(rc, 0)

    wb.save(workflow_Excel_path)
    res_data = ''
    if getExtractDataPref() == True:
        if getPrefSplitpageByLogo():

            res_data =convertReadableFormat(resultPath, pathforclientdirs, "DataProcess_56B")
        else:


            res_data= convertReadableFormat(rootPath, pathforclientdirs, "DataProcess_56B")
    return res_data





# --------------- End of Generate new files list to be processed  -----


resofestimate = ""


def splitpagesbyAssignees(pathforclientdirs, clientdirname, resultPath):
    for root, dirs, files in os.walk(pathforclientdirs):
        makddirectory(clientdirname)

        filecount = 0
        # fileName = os.path.basename(clientdirname).replace("/","-")
        for file in files:
            if file.endswith(".pdf") or file.endswith(".PDF"):

                filecount = filecount + 1

                shutil.copy(os.path.join(root, file), clientdirname)

                # To split the PDFs by pages
                reader = PdfFileReader(os.path.join(clientdirname, file))
                cnt = reader.getNumPages()
                print("Counttt")
                print(cnt)

                countformerge = 0
                temp = 0

                # get all pages text (After image conversion)
                for i in range(cnt):

                    found = False

                    writer = PdfFileWriter()
                    writer.addPage(reader.getPage(i))

                    pdf_bytes = io.BytesIO()
                    writer.write(pdf_bytes)
                    pdf_bytes.seek(0)

                    img = Image(file=pdf_bytes, resolution=600)
                    img.save(filename=os.path.join(clientdirname, file) + str(i) + "-template-target.png")

                    image = cv2.imread(os.path.join(clientdirname, file) + str(i) + "-template-target.png")

                    for image_to_match in os.listdir(os.path.join(cwd + "\\Notice-header-templates")):

                        # template = cv2.imread(os.path.join(cwd+"\\Notice-header-templates", "logo5.png"))
                        # print(image_to_match)

                        template = cv2.imread(os.path.join(cwd + "\\Notice-header-templates", image_to_match))
                        # image = cv2.resize(image, (0,0), fx=0.5, fy=0.5)
                        # template = cv2.resize(template, (0,0), fx=0.5, fy=0.5)

                        # print(os.path.join(cwd+"\\Notice-header-templates", "logo1.png"))

                        # Convert to grayscale

                        imageGray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                        imageGray = cv2.resize(imageGray, None, fx=0.5, fy=0.5)
                        kernel = np.ones((2, 2), np.uint8)
                        imageGray = cv2.erode(imageGray, kernel, iterations=3)

                        kernel = np.ones((3, 3), np.uint8)
                        template = cv2.erode(template, kernel, iterations=2)
                        templateGray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)

                        w, h = templateGray.shape[::-1]

                        w1, h1 = imageGray.shape[::-1]

                        # Crop top part for matching logo here
                        crop_img = imageGray[0:0 + 300, 0:0 + w1]
                        cv2.imwrite(os.path.join(clientdirname, file) + str(i) + "-templateGray-cropped.png", crop_img)

                        tobematched = cv2.imread(
                            os.path.join(clientdirname, file) + str(i) + "-templateGray-cropped.png")

                        tobematched = cv2.cvtColor(tobematched, cv2.COLOR_BGR2GRAY)

                        # print(tobematched)
                        res = cv2.matchTemplate(tobematched, templateGray, cv2.TM_CCOEFF_NORMED)
                        threshold = 0.7
                        loc = np.where(res >= threshold)

                        for pt in zip(*loc[::-1]):
                            found = True
                            cv2.rectangle(tobematched, pt, (pt[0] + w, pt[1] + h), (0, 255, 255), 2)

                        # print("-------->>>> result <<<<<--**------")
                        # print(found)
                        # print(i)
                        # print(resultTemplateMatches)
                        # print(template)
                        cv2.imwrite(os.path.join(clientdirname, file) + str(i) + "-template-detected.png", tobematched)
                        cv2.imwrite(os.path.join(clientdirname, file) + str(i) + "-templateGray.png", templateGray)

                        if found:
                            break

                    # print("----------- found ==============")
                    # print(i)
                    # print(found)
                    if found and i > 0:

                        # found = False
                        # merger.append(os.path.join(clientdirname, file),pages=PageRange(0,i))
                        if countformerge == 0:
                            # print("countformerge")
                            # print(i)
                            merger = PdfFileMerger()
                            merger.append(os.path.join(pathforclientdirs, file), pages=(0, i), import_bookmarks=False)
                            merger.write(os.path.join(resultPath, file) + "page-" + str(i) + "split-by-assignee.pdf")
                            temp = i
                            countformerge = countformerge + 1
                            merger.close()
                        else:
                            merger = PdfFileMerger()
                            # print("========== Temp ============")
                            # print(temp)
                            # print(i)
                            merger.append(os.path.join(pathforclientdirs, file), pages=(temp, i),
                                          import_bookmarks=False)
                            merger.write(os.path.join(resultPath, file) + "page-" + str(i) + "split-by-assignee.pdf")
                            temp = i
                            merger.close()
                    if i == cnt - 1:
                        merger = PdfFileMerger()
                        # print("Temp for last index -->")
                        # print(i + 1)
                        # print(temp)

                        # temp+1 changed to temp now due to include template page for last left
                        merger.append(os.path.join(pathforclientdirs, file), pages=(temp, i + 1),
                                      import_bookmarks=False)
                        merger.write(os.path.join(resultPath, file) + "page-" + str(i) + "split-by-assignee.pdf")
                        merger.close()

                        # FileName = os.path.join(clientdirname, file)+"page-"+str(i)+"split-by-assignee.pdf"
                        # with open(FileName, 'wb') as outfile:
                        # writer.write(outfile)


# Generic functions to check and call comparision methods

def convertReadableFormat(resultPath, pathforclientdirs, clientdirname):

    # try:
    wb = load_workbook(workflow_Excel_path)
    sheet3 = wb['Results_56B']
    sheet2 = wb['Labels to be extracted']
    sheetdoc = wb['DocumentTypes']

    lables_tuple = ()
    docs_tuple = ()

    for cell in sheet2['B']:
        # print(cell.value)
        lables_tuple += (cell.value,)

    # Get all labels from Json


    output_json = json.load(open(os.path.join(cwd, "json/IR56B_extract_labels.json")))

    my_dict = {}

    for element in output_json:
        for value in output_json['DocumentLabels']:
            my_dict.update({value['Labelname']: value['Datatype']})

    for cell in sheetdoc['A']:
        # print(cell.value)
        docs_tuple += (cell.value,)

    # print("-------->>>> Inside createDirFiles <<<<<-----")
    #
    # print(pathforclientdirs)

    pagePref = getPagePreferences()

    rowcnt = 1
    print(resultPath)
    for root, dirs, files in os.walk(resultPath):

        makddirectory(clientdirname)

        filecount = 0
        # fileName = os.path.basename(clientdirname).replace("/","-")


        for file in files:
            print("File", root, file)
            if ("56B" in root):

                if file.endswith(".pdf") or file.endswith(".PDF") or file.endswith(".xlsx") or file.endswith(".xls"):
                    # shutil.copy(os.path.join(root, file), os.path.basename(clientdirname).replace(" ","-")+"-"+os.path.basename(sub).replace(" ","-"))
                    try:

                        shutil.copy(os.path.join(root, file), clientdirname)

                        # To split the PDFs by pages
                        reader = PdfFileReader(os.path.join(clientdirname, file))
                        cnt = reader.getNumPages()
                    except:
                        cnt = 0
                        print("E")


                    textforall = ""
                    # get all pages text (After image conversion)


                    immdFolder = str(root).split("\\")
                    # print(immdFolder[len(immdFolder) - 1])
                    parentdirfordoc = immdFolder[len(immdFolder) - 1]
                    # print(dirs)
                    # print(os.path.basename(file))
                    sheet3.cell(row=rowcnt + 1, column=1).value = os.path.basename(file)

                    FileName = os.path.basename(file)

                    for i in range(cnt):
                        writer = PdfFileWriter()
                        writer.addPage(reader.getPage(i))
                        FileName = os.path.join(clientdirname, file) + "page-" + str(i) + ".pdf"

                        with open(FileName, 'wb') as outfile:
                            writer.write(outfile)

                        # --- Process begins here to check the page has scanned or digital
                        imgflag = pdfminer.validateForImage(FileName)


                        pdfminerText = pdfminer.getTextForPage(FileName)

                        # print("=============  pdfminerText ====================")


                        # if imgflag and (len(pdfminerText)<12):

                        if (len(pdfminerText) < 250):
                            # For converting PDF page to image
                            pdf_file = open(os.path.join(cwd, FileName), 'rb')
                            pdf_reader = PdfFileReader(pdf_file)
                            dst_pdf = PyPDF2.PdfFileWriter()
                            dst_pdf.addPage(pdf_reader.getPage(0))

                            pdf_bytes = io.BytesIO()
                            dst_pdf.write(pdf_bytes)
                            pdf_bytes.seek(0)

                            img = Image(file=pdf_bytes, resolution=600)

                            if i == 0:
                                img.save(filename=os.path.join(clientdirname, file) + str(i) + "-temp-first-page.png")
                            else:
                                img.save(filename=os.path.join(clientdirname, file) + str(i) + "-temp-page.png")
                                # text = pytesseract.image_to_string(ImagePIL.open(os.path.join(clientdirname, file)+str(i)+"-temp.png"), lang='eng', config='psm 11 -oem 0 -c tessedit_char_whitelist=01234567890abcdefghijklmnopqrstuvwxyz')

                            # For OCR the image to hocr and then PDF
                            pdf = FPDF()
                            pdf.add_page()
                            pdf.set_font("Arial", size=14)
                            # pdf.cell(200, 10, txt="Welcome to Python!", align="C")
                            pdf.output(os.path.join(clientdirname, file + "temp.pdf"))

                            if i == 0:
                                imgtobprocess = os.path.join(clientdirname, file) + str(i) + "-temp-first-page.png"
                            else:
                                imgtobprocess = os.path.join(clientdirname, file) + str(i) + "-temp-page.png"
                            # croptextarea(imgtobprocess,i,file)

                            # imgtobprocess = os.path.join(clientdirname, file)+str(0)+"-temp-first-page-cropped-text.png"

                            # img333 = cv2.imread(imgtobprocess)
                            # gray = cv2.cvtColor(img333, cv2.COLOR_BGR2GRAY)

                            # set image dpi
                            temp_filename = set_image_dpi(imgtobprocess, i, file)

                            # print(temp_filename)

                            th3 = remove_noise_and_smooth(temp_filename)

                            cv2.imwrite(os.path.join(clientdirname, file) + str(i) + "-smooth-noise-rem.jpg", th3)

                            # skewedfilepath = skewCorrection(os.path.join(clientdirname, file)+str(i)+"-smooth-noise-rem.jpg",i,file)

                            skewedfilepath = os.path.join(clientdirname, file) + str(i) + "-smooth-noise-rem.jpg"

                            imgtobeprocessedafter = skewedfilepath

                            # imgtobeprocessedafter = savepathcropped
                            # config='-oem 1'
                            pdfres = pytesseract.image_to_pdf_or_hocr(imgtobeprocessedafter, config='--oem 1',
                                                                      extension='pdf')

                            # print(pdfres)

                            filePDF = open(os.path.join(clientdirname, file + "temp.pdf"), 'wb')
                            filePDF.write(pdfres)
                            filePDF.close()
                            pdf_file.close()

                            # filePathForTabula = os.path.join(cwd, FileName)

                            filePathForTabula = os.path.join(clientdirname, file + "temp.pdf")

                            # print(filePathForTabula)
                            TabulaExtract.extracttablesGuessTrue(filePathForTabula, filePathForTabula + str(i) + ".csv")
                            TabulaExtract.extracttablesForAll(filePathForTabula, filePathForTabula + str(i) + "normal.csv")
                            TabulaExtract.extracttablesWithStream(filePathForTabula,
                                                                  filePathForTabula + str(i) + "tabcords.csv")

                            filecount = filecount + 1

                            # Tesseract image to string

                            text = pytesseract.image_to_string(imgtobeprocessedafter, lang='eng',
                                                               config='psm 11 -oem 0 -c tessedit_char_whitelist=01234567890abcdefghijklmnopqrstuvwxyz')
                            # print("text for images ......>>>>>>")
                            # print(text.replace('\n', ' ').replace('\r', ''))

                            try:
                                # print("textforallpages here check if incomplete")
                                # # print(str(textforallpages).replace("** - ",""))
                                #
                                # print(text)
                                textstr = str(text).replace("** - ", "")
                                textstr = textstr.replace('"', '')
                                textstr = textstr.replace('|', '')

                                textstr = textstr.replace('_ ______________ __', '')
                                textstr = textstr.replace('_', '')
                                textstr = textstr.replace('  ', ' ')
                                textstr = textstr.replace('  ', ' ')
                                textstr = textstr.replace('Ref.', 'Ref:')
                                textstr = textstr.replace('Réf:', 'Ref:')
                                textstr = textstr.replace('For. ', 'For:')
                                textstr = textstr.replace('For ', 'For:')
                                # textstr = textstr.replace('Result of this Estimate - Tax Refundable _| ______________ __|','Result of this Estimate')

                                textstr = textstr.replace('For.', 'For:')
                                textstr = textstr.replace('Réf:', 'Ref:')
                                textstr = textstr.replace('Ref.', 'Ref:')
                                # need to place logic for custom cleaning for data
                                textstr = textstr.replace('201 7', '2017')
                                textstr = textstr.replace('£1,8§58.09', '£1,8958.09')
                                textforall = textstr.replace(' 1 hey owe mvmc f301 00', 'They owe hmrc £301.00')



                                textstrsplittebylines = textforall.split("\n")
                                #                                    print("^^^^^^^^Textstrsplittedbylines^^^^^")
                                # print(textstrsplittebylines)
                                with open(os.path.join(clientdirname, file) + "-txtoutbylines.csv", "w") as csv_file:
                                    writer = csv.writer(csv_file)
                                    for data in textstrsplittebylines:
                                        # print("splitted lines here")
                                        # print(data)
                                        writer.writerow([data])

                                    #                                    if  textstrsplittebylines != "":
                            #                                        with open(os.path.join(clientdirname, file)+"-txtoutbylines.csv", "w") as csv_file:
                            #                                            writer = csv.writer(csv_file)
                            #                                            for data in textstrsplittebylines:
                            #                                                print("splitted lines here")
                            #                                                print(data)
                            #                                                writer.writerow([data])
                            #                                    else:
                            #                                        with open(os.path.join(clientdirname, file)+"-txtoutbylines.csv",'w', encoding='utf-8') as f: f.write(str(textforall))
                            #

                            # with open(os.path.join(clientdirname, file)+"-txtout.csv",'w', encoding='utf-8') as f: f.write(str(textforall))
                            # with open(os.path.join(clientdirname, file)+"-txtoutText.txt",'w', encoding='utf-8') as f: f.write(str(textforall))

                            except Exception as e:
                                # with open(os.path.join(clientdirname, file)+"-txtout.csv",'w', encoding='latin1') as f: f.write(str(textforall).replace("** - ",""))

                                #                                        with open(os.path.join(clientdirname, file)+"-txtoutbylines.csv", "w") as csv_file:
                                #                                            writer = csv.writer(csv_file)
                                #                                            for index in range(len(textstrsplittebylines)):
                                #                                                writer.write(textstrsplittebylines[index])
                                #                                                writer.write("\n")

                                # print("Exception is here write in to text file")
                                # print(str(text))
                                print(e)

                            # Data extraction logic

                            # filePathForTabula = filePathForTabula+str(i)+"tabcords.csv"

                            # ******* Changed by Mani ********
                            filePathForTabula = os.path.join(clientdirname, file) + "-txtoutbylines.csv"

                            # pdfminerText = pdfminer.getTextForPage(FileName)
                            # Extract information page by page

                            extractLabelValuePairs(filePathForTabula, sheet3, rowcnt, lables_tuple, docs_tuple, textforall,
                                                   pdfminerText, "Scanned", my_dict, FileName, parentdirfordoc)

                            # ---  For extract tables using camlot -----
                            filePathForPDFGenerated = os.path.join(clientdirname, file + "temp.pdf")
                            # extractTables(filePathForPDFGenerated,rowcnt)

                            rowcnt = rowcnt + 1
                            if pagePref == "First page only":
                                if i == 0:
                                    break

                        else:

                            filePathForTabula = os.path.join(clientdirname, file + "digital.pdf")
                            TabulaExtract.extracttablesWithStream(FileName, filePathForTabula + str(i) + ".csv")
                            TabulaExtract.extracttablesGuessTrue(FileName,
                                                                 filePathForTabula + str(i) + "tables-digital.csv")

                            resfilefromtabula = filePathForTabula + str(i) + ".csv"

                            with open(filePathForTabula + str(i) + ".csv") as file:
                                data = file.read().replace("oo.oo", "00.00")
                                data = data.replace("A mount", "Amount")

                            with open(filePathForTabula + str(i) + ".csv", 'w') as file:
                                file.write(data)

                            extractLabelValuePairs(resfilefromtabula, sheet3, rowcnt, lables_tuple, docs_tuple, textforall,
                                                   pdfminerText, "Digital", my_dict, FileName, parentdirfordoc)

                            filePathForPDFGenerated = FileName
                            # extractTables(filePathForPDFGenerated,rowcnt)

                            rowcnt = rowcnt + 1
                            if pagePref == "First page only":
                                if i == 0:
                                    break
    # Delete dirs/files created after processed



    wb.save(workflow_Excel_path)

    # # For cutom data cleaning for MTR
    response_56B= customDataCleaningForIR56B(workflow_Excel_path, cwd)
    return response_56B


def extractLabelValuePairs(filePathForTabula, sheet3, rowcnt, lables_tuple, docs_tuple, textforall, pdfminerText, mode,
                           my_dict, fileName, parentdirfordoc):
    # Document classification using sentence similarity

    doc_classify = ""
    doc_classify_by_nlp = ""

    RelatedText = ""
    colNum = 0

    personTextMatches = ""
    # additionalEntities = ""

    # # Check Tuple here by loop
    # print("-----  Check Tuple here by loop >>>>>>>>>")
    # for var in lables_tuple[1:]:
    #     print (lables_tuple.index(var), var)

    ResutcolsForExcel = pd.DataFrame(columns=list(my_dict.keys()))
    ResutcolsForExcel.insert(0, "Doc Name", "")

    # -----------------
    colcountforexcelheader = 1
    for col_num, value in enumerate(ResutcolsForExcel.columns.values):
        sheet3.cell(row=1, column=col_num + 1).value = value
        colcountforexcelheader = col_num


    if mode == "Scanned":
        strcleaned = cleanText(textforall)
    #        print("------------ Textforall =====")
    #        print(mode)
    #        print(strcleaned)

    elif mode == "Digital":
        strcleaned = cleanText(pdfminerText)
        # print("------------ Textforall =====")
        # print(mode)
        # print(strcleaned)

    with open(filePathForTabula, newline='', encoding='ISO-8859-1'):
        # reader = csv.reader(f)
        # ncol=len(next(reader))

        file_handle = open(filePathForTabula, "r")

        # print("Total columns")
        # print(ncol)
        # readerfornextline = csv.reader(f)

        col = 1

        sheet3.cell(row=rowcnt + 1, column=1).value = fileName.split("\\")[1]

        for labelvar, fieldtype in my_dict.items():
            # print(labelvar, ":", fieldtype)
            # for labelvar in lables_tuple[1:]:

            file_handle.seek(0)
            col = col + 1
            colNum = col
            # print("outside field --->")
            # print(var)

            brflag = False
            # starFlag = False
            reader3 = csv.reader(file_handle)
            for row in reader3:
                field = ""
                # for field in row:
                field = ' '.join(row)

                field = field.replace("Employee’s Tax File No. with this department :","Employee's Tax File No. with this department :")
                field = field.replace("Employer’s File No. :", "Employer's File No.:")
                field = field.replace("Employer's File No.:",  "Employer's File No.:")
                field = field.replace("Employer File No.", "Employer's File No.:")
                field = field.replace("|", "")
                #incaseo of special characters


                field = field.replace("Residential address:", "Residential address_person:")
                field = field.replace("Name of Employee or Pensioner:", "Given name in Full:")
                field = field.replace("Sex(M=Male,F=Female):", "Sex (M=Male, F=Female):")
                field = field.replace("7. Residential Address:", "Residential address_person:")
                field = field.replace("7. Residential address :", "Residential address_person:")
                field = field.replace("8. Postal Address:", "Postal Address (if different from 7 above):")
                field = field.replace("10. Period of employment for the year:","Period of employment for the year from 1 April 2018 to 31 March 2019:")
                field = field.replace("Salaries Tax paid by Employer (Overseas IIT)", "Salaries Tax Paid by Employer")
                field = field.replace("Nature: Housing Allowances", "Nature1: Cash / housing allowance")
                field = field.replace("Nature: Other Allowances", "Nature2: Other allowance / award")
                field = field.replace("Name of the overseas company:", "Name of the non-Hong Kong company:")
                field = field.replace("company (0=No, 1=Yes)", "by a non-Hong Kong company: (0=No, 1=Yes)")
                field = field.replace("Nature: Cash / housing allowance", "Nature1: Cash / housing allowance")
                field = field.replace("Nature: Other allowance / award", "Nature2: Other allowance / award")
                field = field.replace("Nature: Foreign Individual Income Tax paid by Employer", "Nature3: Foreign Individual Income Tax paid by Employer")
                field = field.replace("Nature: Foreign IIT paid by Employer / staff benefits","Nature3: Foreign Individual Income Tax paid by Employer")
                field = field.replace("N ature: Foreign IIT paid by Employer / staff benefits","Nature3: Foreign Individual Income Tax paid by Employer")
                field = field.replace("P eriod of employment for the year from 1 April 2018 to 31 March 2019:",
                                      "Period of employment for the year from 1 April 2018 to 31 March 2019:")
                field = field.replace("(a) HK. Identity Card Number:", "H.K. Identity Card Number:")
                field = field.replace("§", "")
                field = field.replace("Sex (M-=Male, F=Female):", "Sex (M=Male, F=Female):")



                # field = field.replace("The nature of 1st other rewards", "Nature1: The nature of 1st other rewards")
                # field = field.replace("The nature of 2nd other rewards","Nature2: The nature of 2nd other rewards")
                # field = field.replace("The nature of 3rd other rewards","Nature3: The nature of 3rd other rewards")

                #                    if "**" in labelvar.strip():
                #                        starFlag = True
                #                        labelvar = labelvar.replace("**","")
                # print(">>>>>>>>>>>")
                if labelvar.strip().lower() in field.lower():
                    # print("----------++++ var ++++========>>>")
                    # print(var)
                    # print(field)
                    #                        taxref = field
                    #                        taxref = taxref.lower().split(labelvar.lower().strip())
                    #
                    if fieldtype == "Money":
                        taxref = field
                        # print(taxref)

                    else:
                        taxref = field

                        #                            taxref = "Your refrence sdsd owes ggg"
                        #                            labelvar = "hmrc owes them"
                        if labelvar.lower() == 'our ref' or labelvar.lower() == 'our reference':
                            # print("dfdfdfd")
                            if 'your ref' not in taxref.lower():
                                # print("Inside ....")

                                taxref = taxref.lower().split(labelvar.lower().strip())
                        else:
                            taxref = taxref.lower().split(labelvar.lower().strip())

                    labelvar = labelvar.strip()

                    # print(taxref[1])
                    if fieldtype == "Money":
                        fieldtext = _nlp(taxref)
                    else:
                        fieldtext = _nlp(taxref[1].strip())

                    resfromnlp = GetLabelValuesBYdTypes(labelvar, fieldtext, fieldtype)

                    # if len(taxref) > 2:
                    # sheet3.cell(row=rowcnt+1, column=col).value = taxref[1] + taxref[2]
                    # else:
                    # sheet3.cell(row=rowcnt+1, column=col).value = taxref[1]
                    # print("================resfromnlp-----========")
                    # print(resfromnlp)
                    try:
                        sheet3.cell(row=rowcnt + 1, column=col).value = resfromnlp

                    except Exception as e:
                        if type(resfromnlp) is list:
                            # print("######Inside Exception Block##############")
                            # print(resfromnlp)
                            sheet3.cell(row=rowcnt + 1, column=col).value = ""
                        # print(e)
                        # Check for document classifications
                        # Check if values is empty or not
                    varval = sheet3.cell(row=rowcnt + 1, column=col).value
                    address_list =[]
                    if col==42 and varval != None:
                        brflag = False
                        continue

                    if varval != None and varval.strip() != "":
                        brflag = True
                        break
                    elif len(varval.strip()) != 0:
                        brflag = True
                        break

                if brflag:
                    brflag = True
                    break

            #                if brflag:
            #                    brflag = False
            #                    break



    strcleaned = strcleaned.replace(";", "")
    strcleaned = strcleaned.replace("~", "")
    strcleaned = strcleaned.replace("!", "")


# Spelling correction here

def wordMerger(doc):
    try:
        cnt = 0
        with doc.retokenize() as retokenizer:
            for token in doc:
                if len(token.text) == 1:
                    if token.dep_ == "compound":

                        retokenizer.merge(doc[cnt - 1:cnt + 1], attrs={})
                cnt = cnt + 1
    except Exception as e:
        print(e)
    return doc


# Clean text

def cleanText(text):
    text = text.strip().replace("\n", " ").replace("\r", " ").replace("Nature: HOTEL","Nature_hotel:HOTEL")
    #    text = re.sub(r"\n+", " ", text)
    text = re.sub(' +', ' ', text)
    text = text.replace("MR", "Mr")
    text = text.replace("MRS", "Mrs")
    text = text.replace("Sex(M=Male,F=Female):", "Sex (M=Male, F=Female):")
    text = text.replace("Residential address:", "Residential address_person:")
    text = text.replace("Name of Employee or Pensioner:", "Given name in Full:")
    text = text.replace("7. Residential Address:", "Residential address_person:")
    text = text.replace("7. Residential address :", "Residential address_person:")
    text = text.replace("8. Postal Address:", "Postal Address (if different from 7 above):")
    text = text.replace("10. Period of employment for the year:Period of employment for the year:", "Period of employment for the year from 1 April 2018 to 31 March 2019:")
    text = text.replace("Salaries Tax paid by Employer (Overseas IIT)", "Salaries Tax Paid by Employer")
    text = text.replace("Name of the overseas company:", "Name of the non-Hong Kong company:")
    text = text.replace("company (0=No, 1=Yes)", "by a non-Hong Kong company: (0=No, 1=Yes)")

    text = text.replace("Nature: Cash / housing allowance", "Nature1: Cash / housing allowance")
    text = text.replace("Nature: Other allowance / award", "Nature2: Other allowance / award")
    text = text.replace("Nature: Foreign Individual Income Tax paid by Employer", "Nature3: Foreign Individual Income Tax paid by Employer")
    text = text.replace("Nature: Foreign IIT paid by Employer / staff benefits","Nature3: Foreign Individual Income Tax paid by Employer")
    text = text.replace("N ature: Foreign IIT paid by Employer / staff benefits","Nature3: Foreign Individual Income Tax paid by Employer")
    text = text.replace("(a) HK. Identity Card Number:", "H.K. Identity Card Number:")
    text = text.replace("§", "")
    # print(text)

    return text


# Compute Similarity

# doc1 = "La e tax return Notice of penalty asse sment ERNST & YOUNG LLP"
# doc2 = "Self Assessment: late tax return Notice of penalty assessment"

# doc1 = _nlp(doc1)
# doc2 = _nlp(doc2)

def compute_similarity(doc1, doc2):
    return (doc1.similarity(doc2))


# compute_similarity(doc1, doc2)


# Get particular entities

def getORGintext(doc):
    company = []
    orgent = ""
    for ent in doc.ents:

        # if ent.label_ == 'ORG' and ent.text not in company:
        if ent.label_ == 'ORG':
            company.append(ent.text)
            orgent = orgent + ", " + ent.text

    # print("---------- ORG entity --------------")
    # print(company)

    return orgent


# Include Person Title


# Expand person entity with MR, MRs etc

# doc = _nlp("You or your adviser told us that for the tax years 2016-2017 and 2017-2018 and 2015-2017 no longer need to send us Self")


def getDateintext(doc):
    dateEntity = ""
    for ent in doc.ents:
        if ent.label_ == 'DATE':
            # print("In Date ===>>> ENT ==>")
            # print(ent.text)

            if dateEntity != "":
                dateEntity = dateEntity + " and " + ent.text
            #                print(dateEntity)
            else:
                dateEntity = ent.text
    # print("---------- Person entity --------------")
    # print(perent)
    if not hasNumbersinToken(dateEntity):
        dateEntity = ""

    if 'and' in dateEntity:
        dateEntity_splitted = dateEntity.split("and")
        if len(dateEntity_splitted) > 2:
            dateEntity = str(dateEntity_splitted[0]) + "and" + str(dateEntity_splitted[1])
        else:
            dateEntity = dateEntity
    return dateEntity


# getDateintext(doc)


def getCardinalintext(doc):
    cardinalEntity = ""
    for ent in doc.ents:
        if ent.label_ == 'CARDINAL':
            cardinalEntity = ent.text

    return cardinalEntity


# doc = "IAmount due by 22 Mar 18 £2,541 .28"


def getMoneyintext(doc, label):
    moneyEntity = ""

    moneyEntity = extractMoneyByCustomModel(label, doc)


    if moneyEntity == None or moneyEntity == "":
        docx = _nlp(doc)
        for ent in docx.ents:
            if ent.label_ == 'MONEY':
                moneyEntity = ent.text
    print(" in second moneyEntity")
    print(moneyEntity)

    if moneyEntity == None or moneyEntity == "":
        for ent in docx.ents:
            if ent.label_ == 'CARDINAL':
                moneyEntity = ent.text

    # print(" in third moneyEntity")
    # print(moneyEntity)
    # Additional rules for amount
    if moneyEntity == None or moneyEntity == "":
        moneyEntity = str(doc).split(label)
        # print(moneyEntity)
        if len(moneyEntity) > 0:
            if hasNumbersinToken(moneyEntity[1]):
                moneyEntity = moneyEntity[1]

    # print(" in last  moneyEntity")
    #
    # print(type(moneyEntity))
    if type(moneyEntity) == list:
        moneyEntity = ""
    return moneyEntity


# getMoneyintext(doc,"Money")

def getStringintext(doc):
    return str(doc)


# For extract tables using camlot

def extractTables(PDFpage, cnt):
    resultPath = getResultPathDetails()
    # print(os.path.join(cwd, PDFpage))

    # flavor='stream'

    try:
        # tables = camelot.read_pdf(os.path.join(cwd,PDFpage), flavor='stream')

        PDFpagename = PDFpage.replace("DataProcess_56B\\", "")
        exportpath = os.path.join(resultPath, PDFpagename + str(cnt) + ".csv")
        #
        # print("exportpath-------------------------->")
        # print(exportpath)
        # tables.export(exportpath, f='csv', compress=False) # json, excel, html
    except Exception as e:
        print(e)




def hasNumbersinToken(inputString):
    return any(char.isdigit() for char in inputString)


def getAlphaNumericFromText(textline, labeltext):
    doc = textline
    # search_doc_no_stop_words = _nlp(' '.join([str(t) for t in doc if not t.is_stop]))
    # search_doc_no_stop_words = _nlp(' '.join([str(t) for t in search_doc_no_stop_words if not len(t)==1]))

    alphanum = ""

    doc = str(doc).replace("-", "")
    doc = str(doc).replace("_", "")
    # print("In AlphaNumeric here --->")
    # print(doc)
    docx = _nlp(doc)



    for token in docx:



        if hasNumbersinToken(
                token.text) or token.dep_ == "subtok" or token.dep_ == "ROOT" or token.dep_ == "appos" or token.dep_ == "nummod" or token.dep_ == "npadvmod" or token.dep_ == "pobj":
            alphanum = alphanum + token.text

    #        for ent in tokendoc.ents:
    #            print(ent.text, ent.start_char, ent.end_char, ent.label_)
    #             alphanum =  ent.text
    if not hasNumbersinToken(alphanum):
        alphanum = ""
    else:
        alphanum = alphanum.replace("_", "")
        alphanum = alphanum.replace(":", "")
    # print(alphanum)
    return alphanum



nlp3 = spacy.load(spacyCustomModelpath)


def extractEntitiesByModel(labelvar, textforall):
    doc2 = nlp3(textforall.lower())

    labelval = ""
    for ent in doc2.ents:
        # print(ent.label_, ent.text)
        if labelvar.lower() in ent.label_.lower():
            # print(ent.label_, ent.text)
            labelval = labelval + ent.text
    #
    # print(labelval)
    return labelval


def extractMoneyByCustomModel(labelvar, textforall):
    doc2 = nlp3(textforall.lower())

    # print("In custom entities model text to test")
    # print(labelvar)

    if labelvar == "Money":
        print("In custom entities model text to test")
        # print(doc2)

    labelval = ""
    for ent in doc2.ents:
        labelval = labelval + ent.text
    # print(labelval)
    return labelval


# extractMoneyByCustomModel("Amount due by","IAmount due by 22 Mar 18 100.00 1")


def extractAdditionalEntitiesByModel(textforall):

    doc2 = nlp3(str(textforall))

    labelval = ""
    for ent in doc2.ents:
        # print(ent.label_, ent.text)
        labelval = labelval + ", " + ent.label_ + ":" + ent.text
    # print(labelval)
    return labelval


# ----------------  testing ----------------

# main callable method


def GetLabelValuesBYdTypes(labeltext, textline, fieldtype):


    result = ""

    if fieldtype == "Number":

        result = getAlphaNumericFromText(textline, labeltext)

    if fieldtype == "Date":

        result = getDateintext(textline)

    if fieldtype == "Money":

        result = getMoneyintext(textline, labeltext)


    if fieldtype == "String":

        result = getStringintext(textline)

    return result



wb = load_workbook(workflow_Excel_path)
sheet3 = wb['Results_56B']





